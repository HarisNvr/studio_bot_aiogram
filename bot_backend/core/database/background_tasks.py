from asyncio import to_thread
from datetime import datetime, timedelta
from os import remove
from pathlib import Path

from aiogram.types import Message
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pandas import DataFrame
from sqlalchemy import delete, select, update, insert

from core.components.settings import TZ, BASE_DIR
from core.database.db_connection import async_session_maker
from core.database.models import UserMessage, User
from core.utils.path_builder import get_file
from core.utils.tg_channel import sub_check


async def morning_routine():
    """
    Delete old message IDs from the DB. Telegram's policy doesn't allow bots
    to delete messages that are older than 48 hours.

    :return: None
    """

    threshold = datetime.now(TZ) - timedelta(hours=48)
    stmt = delete(UserMessage).where(
        UserMessage.message_date < threshold
    )

    async with async_session_maker() as session:
        await session.execute(stmt)
        await session.commit()


async def get_user_id(message: Message) -> int | None:
    """
    Retrieves the user's primary key from the database using the chat id.

    :param message: The message sent by the user.
    :return: User's primary key - ID or None
    """

    stmt = select(User).where(User.chat_id == message.chat.id)

    async with async_session_maker() as session:
        result = await session.execute(stmt)
        user = result.scalar_one_or_none()

        if user:
            return user.id


async def record_message_id_to_db(*messages: Message):
    """
    Record message id's to DB, for 'clean' func.

    :param messages: The message sent by the user.
    :return: None
    """

    stmt_query = [
        insert(UserMessage).values(
            user_id=await get_user_id(message),
            message_id=message.message_id
        ) for message in messages
    ]

    async with async_session_maker() as session:
        for stmt in stmt_query:
            await session.execute(stmt)
        await session.commit()


async def write_user(message: Message):
    """
    Record user's data to DB.

    :param message: The message sent by the user.
    :return: None
    """

    is_subscribed = await sub_check(message.chat.id)

    stmt = insert(User).values(
        chat_id=message.chat.id,
        username=message.from_user.username,
        user_first_name=message.from_user.first_name,
        is_subscribed=is_subscribed
    )

    async with async_session_maker() as session:
        await session.execute(stmt)
        await session.commit()


async def update_user(message: Message):
    """
    Update user's data in DB.

    :param message: The message sent by the user.
    :return: None
    """

    is_subscribed = await sub_check(message.chat.id)

    stmt = update(User).where(User.chat_id == message.chat.id).values(
        username=message.from_user.username,
        user_first_name=message.from_user.first_name,
        is_subscribed=is_subscribed
    )

    async with async_session_maker() as session:
        await session.execute(stmt)
        await session.commit()


async def get_users(message: Message = None) -> None:
    """
    Exports the Users table to an Excel file, sends it to the user through
    Aiogram, and deletes the file after sending.

    :param message: Telegram user message.
    :return: None
    """

    async with async_session_maker() as session:
        result = await session.execute(select(User).order_by(User.id))
        users = result.scalars().all()

        data = [
            {
                'ID': user.id,
                'ID чата': user.chat_id,
                'Никнейм': user.username or '-',
                'Имя пользователя': user.user_first_name,
                'Подписан на ТГ': 'ДА' if user.is_subscribed else 'НЕТ'
            }
            for user in users
        ]

        df = DataFrame(data)
        now = datetime.now().strftime('%d-%m-%Y')

        file_name = f'Таблица пользователей от {now}.xlsx'
        file_path = BASE_DIR / file_name

        df.to_excel(file_name, index=False, engine='openpyxl')

        ws = load_workbook(file_name).active

        for col in ws.iter_cols():
            max_length = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)].width = max_length + 2
            for cell in col:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

        wb = ws.parent
        wb.save(file_name)
        wb.close()

        await message.answer_document(
            document=get_file(
                file_name=file_name,
                directory=BASE_DIR
            )
        )

        await to_thread(remove, file_path)
